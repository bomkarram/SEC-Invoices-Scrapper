from private_data import locations
from my_selenium import Toolkit
from invoice import Invoice

# for excel manipulation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, colors

# progress bar
from tqdm import tqdm

if __name__ == '__main__':

    # initiate driver
    toolkit = Toolkit()
    driver = toolkit.driver

    wb = Workbook()
    # remove default worksheet
    ws = wb.active
    wb.remove(ws)

    total_accounts = sum(map(lambda account_list: len(account_list), locations.values()))
    pbar = tqdm(total=100)

    for place, accounts in locations.items():
        # setup excel worksheet
        ws = wb.create_sheet(place)
        ws.append(['الوصف', 'رقم الحساب'])

        for account_description, account_number in accounts:

            # for ...:
            #    While True:
            #       try:
            #           ...
            #           break
            #       except...:
            #           ...
            #           continue
            # explanation: don't leave current loop unless it succeed

            while True:
                try:
                    # get invoice of this account
                    invoice = Invoice(toolkit, driver, account_number)

                    # insert result to excel file
                    ws.append([account_description, account_number, invoice.total_amount, invoice.payment])
                    ws['C1'] = invoice.invoice_date

                    # # Debug
                    # print([account_description, account_number, invoice.total_amount, invoice.payment])

                    if float(invoice.total_amount) < 300:
                        ws['C' + str(ws.max_row)].fill = PatternFill("solid", fgColor=colors.YELLOW)
                    elif float(invoice.total_amount) < 500:
                        ws['C' + str(ws.max_row)].fill = PatternFill("solid", fgColor=colors.GREEN)
                    elif float(invoice.total_amount) < 1000:
                        ws['C' + str(ws.max_row)].fill = PatternFill("solid", fgColor=colors.RED)
                        ws['C' + str(ws.max_row)].font = Font(color=colors.WHITE)
                    else:
                        ws['C' + str(ws.max_row)].fill = PatternFill("solid", fgColor=colors.BLACK)
                        ws['C' + str(ws.max_row)].font = Font(color=colors.WHITE)

                    # push progress bar one step
                    pbar.update(1)

                    # # Debug
                    # print(driver.page_source)
                    # driver.save_screenshot(account_number + '.png')

                    break

                except Exception as e:
                    # # Debug
                    # print(e)
                    # driver.save_screenshot(account_number + '_error.png')
                    # print(driver.page_source)

                    driver.quit()
                    toolkit = Toolkit()
                    driver = toolkit.driver
                    continue

    wb.save("invoices.xlsx")
    driver.quit()
